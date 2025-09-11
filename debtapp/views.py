# =========================
# Standard Library Imports
# =========================
from datetime import timedelta, date, datetime
from io import BytesIO
from decimal import Decimal
from datetime import date, datetime, time

# =========================
# Django Imports
# =========================
from django.contrib import messages
from django.contrib import messages
from django.contrib import messages
from django.contrib import messages
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm
from django.db import transaction as db_transaction
from django.db import transaction as db_transaction
from django.db.models import Sum
from django.db.models import Count, Sum, F, Value, Q
from django.db.models import OuterRef, Subquery, DecimalField
from django.db.models.functions import Coalesce
from django.http import HttpResponse, HttpResponseBadRequest
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.utils.html import strip_tags
from django.views import View
from django.views.decorators.cache import never_cache
from django.views.decorators.cache import never_cache

# =========================
# Third-Party Imports
# =========================
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# =========================
# Email Imports
# =========================
from django.core.mail import send_mail, EmailMessage
from django.core.mail import EmailMultiAlternatives

# =========================
# Local App Imports
# =========================
from .forms import UserRegisterForm, TransactionSearchForm, DebtorForm, TransactionForm
from .models import Debtor, Transaction, CustomUser
from .models import Debtor, Transaction

# =========================
# Constants / Helpers
# =========================
ZERO = Value(0, output_field=DecimalField(max_digits=12, decimal_places=2))


# =========================
# Profile View
# =========================
@login_required
def creditor_detail(request):
    creditor = request.user
    return render(request, 'creditor_detail.html', {'creditor': creditor})


# =========================
# Register View
# =========================
def register(request):
    if request.method == 'POST':
        form = UserRegisterForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('login')
    else:
        form = UserRegisterForm()
    return render(request, 'registration/register.html', {'form': form})


# =========================
# Custom Redirect View
# =========================
@login_required
def custom_redirect_view(request):
    if request.user.is_staff:
        return redirect('admin_dashboard')  # Redirect to the admin dashboard
    else:
        return redirect('user_dashboard')  # Redirect to the user dashboard


# =========================
# Login View
# =========================
def login_view(request):
    if request.method == "POST":
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)  # Perform the login action

            # Check user roles based on your custom fields (is_admin, is_user)
            if user.is_admin:  # Admin user
                return redirect('admin_dashboard')  # Redirect to admin dashboard
            elif user.is_user:  # Regular user
                return redirect('user_dashboard')  # Redirect to user dashboard

        else:
            messages.error(request, "Invalid username or password. Please try again.")
    else:
        form = AuthenticationForm()

    return render(request, 'login.html', {'form': form})


# =========================
# Dashboard View
# =========================
@login_required
@never_cache
def dashboard(request):
    base = Debtor.objects.filter(created_by=request.user)

    # If FK uses related_name='transactions', this is correct. Otherwise use 'transaction__...'
    annotated = base.annotate(
        total_debit=Coalesce(Sum('transactions__debit_amount'), ZERO),
        total_credit=Coalesce(Sum('transactions__credit_amount'), ZERO),
        remaining_debt=F('total_debit') - F('total_credit'),  # <-- not 'current_debt'
    )

    active_debtors = annotated.filter(debtor_status='active', is_delete=False)
    recovered_debtors = annotated.filter(debtor_status='recovered', is_delete=False)
    deleted_debtors = annotated.filter(is_delete=True)

    totals = Transaction.objects.filter(debtor__created_by=request.user).aggregate(
        total_debit_amount=Coalesce(Sum('debit_amount'), ZERO),
        total_credit_amount=Coalesce(Sum('credit_amount'), ZERO),
    )
    total_debit_amount = totals['total_debit_amount']
    total_credit_amount = totals['total_credit_amount']
    total_current_debt = total_debit_amount - total_credit_amount

    context = {
        'debtors': annotated,
        'active_debtors': active_debtors,
        'recovered_debtors': recovered_debtors,
        'deleted_debtors': deleted_debtors,

        'total_debtors_no': base.count(),
        'active_debtors_no': base.filter(debtor_status='active', is_delete=False).count(),
        'recovered_debtors_no': base.filter(debtor_status='recovered', is_delete=False).count(),
        'deleted_debtors_no': base.filter(is_delete=True).count(),

        'total_debt_amount': total_debit_amount,
        'total_current_debt': total_current_debt,
        'total_recovered_debt': total_credit_amount,
    }
    return render(request, 'dashboard.html', context)


# =========================
# Logout View
# =========================
def log_out(request):
    logout(request)
    return redirect('login')


# =========================
# Add Debtor
# =========================
@login_required
@never_cache
def add_debtor(request):
    debtor_limit = 50
    debtor_count = Debtor.objects.filter(is_delete=False, created_by=request.user).count()
    if debtor_count >= debtor_limit:
        return render(request, 'limit_exceeded.html')

    if request.method == 'POST':
        form = DebtorForm(request.POST, request.FILES)
        if form.is_valid():
            debtor = form.save(commit=False)
            debtor.created_by = request.user
            debtor.save()

            Transaction.objects.create(
                debtor=debtor,
                tran_type='debit',
                tran_amount=debtor.initial_debt,
                debit_amount=debtor.initial_debt,
                credit_amount=0,
                current_debt=debtor.initial_debt
            )

            # send email
            # subject = 'Add new debtor'
            # message = 'New debtor has been created successfully'
            # from_email = 'raidambar12@gmail.com'
            # recipient_list = ['rairoshan11803@gmail.com']
            #
            # msg_email = EmailMessage(subject, message, from_email, recipient_list)
            # msg_email.send(fail_silently=False)

            # Email Setup with more Control
            subject = "Add new Debtor"
            html_body = render_to_string("emails/email_to_send.html", {"debtor": debtor})
            text_body = strip_tags(html_body)
            email = EmailMultiAlternatives(
                subject=subject,
                body=text_body,
                from_email="no-reply@example.com",  # uses DEFAULT_FROM_EMAIL
                to=["rairoshan11803@gmail.com"]
            )
            email.attach_alternative(html_body, "text/html")
            email.send(fail_silently=True)

            messages.success(request, "1 debtor has been successfully created.")
            return redirect('debtor_list')
    else:
        form = DebtorForm()
    return render(request, 'add_debtor.html', {'form': form})


# =========================
# Debtor List
# =========================
@login_required
@never_cache
def debtor_list(request):
    debtors = Debtor.objects.filter(created_by=request.user, is_delete=False)
    return render(request, 'debtor_list.html', {'debtors': debtors})


# =========================
# Edit Debtor
# =========================
@login_required
@never_cache
def debtor_edit(request, debtor_id):
    debtor = get_object_or_404(Debtor, id=debtor_id, is_delete=False)

    # reverse name: use .transactions if you set related_name, else .transaction_set
    txs = debtor.transactions.all()  # or debtor.transactions.all()
    has_activity = txs.count() > 1   # >1 means beyond the opening row

    if request.method == 'POST':
        form = DebtorForm(request.POST, request.FILES, instance=debtor)
        if form.is_valid():
            old_initial = debtor.initial_debt
            obj = form.save(commit=False)

            if has_activity:
                # Lock initial_debt if there are post-opening transactions
                obj.initial_debt = old_initial
                obj.save()
            else:
                # No activity yet → allow changing initial_debt
                obj.save()

                # Keep the opening transaction in sync with the new initial amount
                opening = txs.order_by('id').first()
                if opening:
                    opening.tran_amount = obj.initial_debt
                    opening.debit_amount = obj.initial_debt
                    opening.credit_amount = 0
                    opening.current_debt = obj.initial_debt
                    opening.save(update_fields=[
                        'tran_amount', 'debit_amount', 'credit_amount', 'current_debt'
                    ])

            messages.success(request, "Debtor updated.")
            return redirect('debtor_list')
    else:
        form = DebtorForm(instance=debtor)
        if has_activity:
            form.fields['initial_debt'].disabled = True

    return render(request, 'debtor_edit.html', {'form': form, 'debtor': debtor})


# =========================
# Transaction Search
# =========================
@login_required
@never_cache
def transaction_search(request):
    debtor_qs = Debtor.objects.filter(is_delete=False, created_by=request.user)

    if request.method == 'POST':
        form = TransactionSearchForm(request.POST)
        if form.is_valid():
            debtor_id = form.cleaned_data['debtor_id']
            tran_type = form.cleaned_data['tran_type']
            return redirect(f"{reverse('add_transaction')}?debtor_id={debtor_id}&tran_type={tran_type}")
    else:
        form = TransactionSearchForm()

    return render(request, 'transaction_search.html', {'form': form, 'debtor': debtor_qs})


# =========================
# Add Transaction
# =========================
@login_required
@never_cache
@db_transaction.atomic
def add_transaction(request):
    debtor_id = request.GET.get('debtor_id')
    tran_type = request.GET.get('tran_type')

    if not debtor_id or not tran_type:
        # messages.error(request, "Missing debtor information.")
        return redirect('transaction_search')

    # Lock the debtor row for this transaction
    debtor = get_object_or_404(
        Debtor.objects.select_for_update(),
        debtor_id=debtor_id,
        created_by=request.user,
        is_delete=False,
    )

    if request.method == 'POST':
        form = TransactionForm(request.POST, request.FILES)
        if form.is_valid():
            tran_amount = Decimal(form.cleaned_data['tran_amount'])
            tran_medium = form.cleaned_data['tran_medium']
            tran_desc = form.cleaned_data.get('tran_desc', '').strip()
            tran_voucher = form.cleaned_data.get('tran_voucher')

            # Validation for positive transaction amount
            if tran_amount <= 0:
                messages.error(request, "Transaction amount must be positive.")
                # re-render preserving query params
                return render(
                    request,
                    'add_transaction',
                    {
                        'form': form,
                        'debtor_id': debtor_id,
                        'tran_type': tran_type
                    }
                )

            current_before = debtor.current_debt

            # Compute new balance based on transaction type
            if tran_type == 'debit':
                current_after = current_before + tran_amount
                debit_amount = tran_amount
                credit_amount = Decimal(0)
                debtor.debtor_status = 'active'
            else:  # credit
                if tran_amount > current_before:
                    messages.error(request, "Cannot enter amount greater than current debt.")
                    return render(
                        request,
                        'add_transaction.html',
                        {
                            'form': form,
                            'debtor_id': debtor_id,
                            'tran_type': tran_type
                        }
                    )
                current_after = current_before - tran_amount
                credit_amount = tran_amount
                debit_amount = Decimal(0)
                if current_after == 0:
                    # is_debt_settle = True  # Debt is settled when current debt is zero
                    debtor.debtor_status = 'recovered'

            # Create the transaction record
            Transaction.objects.create(
                debtor=debtor,
                recorded_by=request.user,
                debit_amount=debit_amount,
                credit_amount=credit_amount,
                tran_amount=tran_amount,
                tran_type=tran_type,
                tran_medium=tran_medium,
                tran_desc=tran_desc,
                current_debt=current_after,
                tran_voucher=tran_voucher,
                # is_debt_settle=is_debt_settle,  # Debt settlement flag
            )

            debtor.save(update_fields=['debtor_status'])
            messages.success(request, f"{tran_type.title()} transaction added successfully.")
            return redirect('debtor_list')
    else:
        form = TransactionForm()

    return render(
        request,
        'add_transaction.html',
        {
            'form': form,
            'debtor_id': debtor_id,
            'tran_type': tran_type,
        }
    )


# =========================
# Debtor Detail
# =========================
@login_required
@never_cache
def debtor_detail(request, debtor_id):
    debtor = get_object_or_404(Debtor, id=debtor_id, created_by=request.user)
    transactions = Transaction.objects.filter(debtor=debtor)
    return render(request, 'debtor_detail.html', {'debtor': debtor, 'transactions': transactions})


# =========================
# Voucher View
# =========================
@login_required
@never_cache
def voucher_view(request, pk):
    debtor = Debtor.objects.filter(id=pk, created_by=request.user).first()
    if debtor:
        return render(request, 'voucher_view.html', {'debtor': debtor})

    transaction = get_object_or_404(Transaction, id=pk, debtor__created_by=request.user)
    return render(request, 'voucher_view.html', {'transaction': transaction})


# =========================
# Soft Delete Debtor
# =========================
@login_required
@never_cache
def delete_debtor(request, id):
    debtor = get_object_or_404(Debtor, id=id)
    transactions = Transaction.objects.filter(debtor=debtor)

    if debtor.debtor_status == 'recovered':
        debtor.is_delete = True
        debtor.delete_date = timezone.now()
        debtor.save(update_fields=['is_delete', 'delete_date'])
        messages.success(request, "1 Debtor has been successfully deleted.")
    else:
        messages.error(request, "Debt is still pending. Cannot delete debtor.")

    return redirect('debtor_list')


# =========================
# Recycle Bin for Debtors
# =========================
@login_required
@never_cache
def recycle_debtor(request):
    debtors = Debtor.objects.filter(created_by=request.user, is_delete=True)
    threshold = timezone.now() - timedelta(days=20)
    expired_debtors = Debtor.objects.filter(delete_date__lt=threshold)

    if expired_debtors.exists():
        expired_debtors.delete()

    return render(request, 'recycle_debtor.html', {'debtors': debtors})


# =========================
# Restore Debtor
# =========================
@login_required
@never_cache
def restore_debtor(request, id):
    debtor_to_restore = Debtor.objects.get(id=id)
    debtor_to_restore.is_delete = False
    debtor_to_restore.save()
    messages.success(request, "1 Debtor restored successfully.")
    return redirect('debtor_list')


# =========================
# Hard Delete Debtor
# =========================
@login_required
@never_cache
def hard_delete_debtor(request, id):
    if request.method != 'POST':
        messages.error(request, "Invalid request method.")
        return redirect('debtor_list')

    debtor = get_object_or_404(Debtor, id=id, created_by=request.user)
    if debtor.debtor_status == 'recovered':
        debtor.delete()
        messages.success(request, "Debtor permanently deleted.")
    else:
        messages.error(request, "Cannot delete debtor with pending debt.")
    return redirect('debtor_list')


# =========================
# Reports (Page Shell)
# =========================
@login_required
def reports(request):
    return render(request, 'reports.html')


# =========================
# Summary Report (User)
# =========================
def summary_details(request):
    qs = Debtor.objects.filter(created_by=request.user)

    total_debtors = qs.count()
    active_debtors = qs.filter(debtor_status='active', is_delete=False).count()
    recovered_debtors = qs.filter(debtor_status='recovered').count()
    deleted_debtors = qs.filter(is_delete=True).count()

    # calculating totals
    totals = Transaction.objects.filter(debtor__created_by=request.user).aggregate(
        total_debit_amount=Sum('debit_amount'),
        total_credit_amount=Sum('credit_amount')
    )
    total_debit_amount = totals['total_debit_amount'] or 0
    total_credit_amount = totals['total_credit_amount'] or 0
    total_current_debt = total_debit_amount - total_credit_amount
    total_recovered_debt = total_credit_amount

    # --- Create workbook ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard Summary"

    # Basic styles
    header_font = Font(bold=True)
    title_font = Font(size=14, bold=True)
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(wrap_text=True)
    fill = PatternFill("solid", fgColor="E8F4FF")
    thin = Side(border_style="thin", color="CCCCCC")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    title = f"Debtors Summary ({timezone.localtime().strftime('%Y-%m-%d, %H:%M')})"
    ws.merge_cells("A1:B1")
    ws["A1"] = title
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    # Headers
    ws["A3"] = "Metric"
    ws["B3"] = "Value"
    ws["A3"].font = header_font
    ws["B3"].font = header_font
    ws["A3"].fill = fill
    ws["B3"].fill = fill
    ws["A3"].border = box
    ws["B3"].border = box
    ws["A3"].alignment = center
    ws["B3"].alignment = center

    # Data rows
    rows = [
        ("Total Debtors", total_debtors),
        ("Active Debtors", active_debtors),
        ("Recovered Debtors", recovered_debtors),
        ("Deleted Debtors", deleted_debtors),
        ("Total Debit Amount", float(total_debit_amount)),
        ("Total Recovered Debt (Credits)", float(total_recovered_debt)),
        ("Total Current Debt", float(total_current_debt)),
    ]

    start_row = 4
    for idx, (label, value) in enumerate(rows, start=start_row):
        ws[f"A{idx}"] = label
        ws[f"B{idx}"] = value
        ws[f"A{idx}"].alignment = wrap
        ws[f"B{idx}"].alignment = center
        ws[f"A{idx}"].border = box
        ws[f"B{idx}"].border = box

    # Number formatting for money-like values (last 3 rows)
    money_rows_start = start_row + 4  # first monetary row index
    for r in range(money_rows_start, money_rows_start + 3):
        ws[f"B{r}"].number_format = '#,##0.00'

    # Column widths
    ws.column_dimensions[get_column_letter(1)].width = 32  # Metric
    ws.column_dimensions[get_column_letter(2)].width = 22  # Value

    # --- Prepare HTTP response ---
    filename = f"dashboard-summary-{timezone.localtime().strftime('%Y%m%d-%H%M')}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


# =========================
# All Debtors Report (User)
# =========================
def all_debtors_xls(request):
    # Annotate each debtor with the current_debt from the most recent transaction (if any)
    latest_current_debt = (
        Transaction.objects
        .filter(debtor=OuterRef('pk'))
        .order_by('-tran_date')
        .values('current_debt')[:1]
    )

    debtors = (
        Debtor.objects
        .filter(created_by=request.user)  # add .filter(is_delete=False) if you want to exclude deleted
        .annotate(current_debt_calc=Subquery(
            latest_current_debt,
            output_field=DecimalField(max_digits=12, decimal_places=2)
        ))
        .order_by('name')
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Debtors"

    # Styling
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    fill = PatternFill("solid", fgColor="E8F4FF")
    thin = Side(border_style="thin", color="CCCCCC")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Name", "Debtor_ID", "Mobile", "Starting Debt", "Current Debt", "Start Date", "Purpose", "Status"]

    # Header row
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.alignment = center
        cell.fill = fill
        cell.border = box

    # Data rows
    row = 2
    for d in debtors:
        current_debt = d.current_debt_calc if d.current_debt_calc is not None else d.total_debt

        ws.cell(row=row, column=1, value=d.name).alignment = left
        ws.cell(row=row, column=2, value=d.debtor_id).alignment = center
        ws.cell(row=row, column=3, value=d.mobile).alignment = center

        c4 = ws.cell(row=row, column=4, value=float(d.initial_debt))
        c5 = ws.cell(row=row, column=5, value=float(current_debt))
        c4.number_format = '#,##0.00'
        c5.number_format = '#,##0.00'
        c4.alignment = center
        c5.alignment = center

        c6 = ws.cell(row=row, column=6, value=d.debt_date)  # DateField -> Excel date
        c6.number_format = 'yyyy-mm-dd'
        c6.alignment = center

        ws.cell(row=row, column=7, value=d.debt_purpose).alignment = left
        ws.cell(row=row, column=8, value=d.get_debtor_status_display()).alignment = center

        # Borders
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = box

        row += 1

    # Column widths (simple auto-fit heuristic)
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        max_len = 0
        for r in range(1, row):
            val = ws.cell(row=r, column=col).value
            max_len = max(max_len, len(str(val)) if val is not None else 0)
        ws.column_dimensions[letter].width = min(max(12, max_len + 2), 40)

    # Response
    ts = timezone.localtime().strftime('%Y%m%d-%H%M')
    filename = f"debtors-detailed-{ts}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# =========================
# Export Transactions (User)
# =========================
@login_required
def debtor_transactions_xls(request):
    debtor_id = request.GET.get("debtor_id")  # e.g., D00001
    if not debtor_id:
        return HttpResponseBadRequest("Missing required parameter: debtor_id")

    debtor = get_object_or_404(Debtor, debtor_id=debtor_id, created_by=request.user)

    txns = (
        Transaction.objects
        .filter(debtor=debtor)
        .select_related("recorded_by")
        .order_by("tran_date", "id")
    )

    # Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{debtor.debtor_id}"

    # Styles
    header_font = Font(bold=True)
    title_font = Font(size=14, bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    fill = PatternFill("solid", fgColor="E8F4FF")
    thin = Side(border_style="thin", color="CCCCCC")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title (no timestamp)
    ws.merge_cells("A1:I1")
    ws["A1"] = f"Transactions for {debtor.name} ({debtor.debtor_id})"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    # Headers
    headers = [
        "Txn ID", "Date", "Type", "Debit", "Credit",
        "Txn Amount", "Current Debt", "Medium", "Description"
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = header_font
        c.alignment = center
        c.fill = fill
        c.border = box

    # Data rows
    row = 4
    total_debit = 0.0
    total_credit = 0.0

    for t in txns:
        ws.cell(row=row, column=1, value=t.tran_id).alignment = center

        # Write date as a plain string -> avoids timezone/aware datetime issues in Excel
        date_str = t.tran_date.strftime("%Y-%m-%d %H:%M")
        ws.cell(row=row, column=2, value=date_str).alignment = center

        ws.cell(row=row, column=3, value=t.tran_type.capitalize()).alignment = center

        c_debit = ws.cell(row=row, column=4, value=float(t.debit_amount))
        c_credit = ws.cell(row=row, column=5, value=float(t.credit_amount))
        c_debit.number_format = "#,##0.00"; c_debit.alignment = center
        c_credit.number_format = "#,##0.00"; c_credit.alignment = center

        ws.cell(row=row, column=6, value=float(t.tran_amount)).number_format = "#,##0.00"
        ws.cell(row=row, column=6).alignment = center

        ws.cell(row=row, column=7, value=float(t.current_debt)).number_format = "#,##0.00"
        ws.cell(row=row, column=7).alignment = center

        ws.cell(
            row=row,
            column=8,
            value=(t.get_tran_medium_display() if hasattr(t, "get_tran_medium_display") else t.tran_medium)
        ).alignment = center

        ws.cell(row=row, column=9, value=t.tran_desc).alignment = left

        # Borders for the row
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = box

        total_debit += float(t.debit_amount)
        total_credit += float(t.credit_amount)
        row += 1

    # Totals row
    ws.cell(row=row, column=3, value="Totals:").font = header_font
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row, column=4, value=total_debit).number_format = "#,##0.00"; ws.cell(row=row, column=4).alignment = center
    ws.cell(row=row, column=5, value=total_credit).number_format = "#,##0.00"; ws.cell(row=row, column=5).alignment = center
    for col in range(1, len(headers) + 1):
        ws.cell(row=row, column=col).border = box

    # Column widths
    widths = [14, 18, 10, 14, 14, 14, 16, 14, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Response (no date/time in filename)
    filename = f"transactions-{debtor.debtor_id}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# =========================
# Admin Dashboard
# =========================
def admin_dashboard(request):
    # Fetch all users (admins and regular users)
    # Total
    users = CustomUser.objects.all()
    debtors = Debtor.objects.all()
    trans = Transaction.objects.all()

    # Total Counts
    total_user_count = CustomUser.objects.count()
    total_debtor_count = Debtor.objects.count()
    total_transaction_count = Transaction.objects.count()

    # Debtors per user
    user_with_debtor_count = (
        CustomUser.objects.annotate(
            debtor_count=Count("debtors", filter=Q(debtors__is_delete=False))
        )
        .values_list("username", "debtor_count").order_by("username")
    )

    # Transactions Per Debtor
    debtor_with_txn = (
        Debtor.objects.filter(is_delete=False)
        .annotate(txn_count=Count("transactions"))
        .values_list("debtor_id", "txn_count").order_by("debtor_id")
    )

    context = {
        'users': users,
        'debtors': debtors,
        'transactions': trans,
        # total Number counts
        'total_user_count': total_user_count,
        'total_debtor_count': total_debtor_count,
        'total_transaction_count': total_transaction_count,
        'user_with_debtor_count': user_with_debtor_count,
        'debtor_with_txn': debtor_with_txn
    }

    return render(request, 'admin1180/admin_dashboard.html', context)


# =========================
# Admin: User (Creditor) Profile
# =========================
@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
@never_cache
def admin_creditor_detail(request, pk):
    User = get_user_model()
    # show the requested user's profile, not the logged-in admin
    creditor = get_object_or_404(User, pk=pk)

    # list this user's active (non-deleted) debtors + txn counts
    debtors = (
        Debtor.objects
        .filter(created_by=creditor, is_delete=False)
        .annotate(txn_count=Count("transactions"))
        .order_by("debtor_id")
    )

    return render(
        request,
        "admin1180/admin_creditor_detail.html",
        {"creditor": creditor, "debtors": debtors},
    )


# =========================
# Admin: Debtor Detail
# =========================
@login_required
@never_cache
def admin_debtor_detail(request, pk):
    # Admins can see any debtor; non-admins can only see their own
    base_qs = Debtor.objects.select_related("created_by").filter(is_delete=False)

    if request.user.is_staff or request.user.is_superuser:
        debtor = get_object_or_404(base_qs, pk=pk)
    else:
        debtor = get_object_or_404(base_qs, pk=pk, created_by=request.user)

    transactions = (
        Transaction.objects
        .filter(debtor=debtor)
        .order_by("tran_date")  # or "-tran_date" if you prefer recent first
    )

    return render(
        request,
        "admin1180/admin_debtor_detail.html",
        {"debtor": debtor, "transactions": transactions},
    )


# =========================
# Admin Profile Detail
# =========================
@staff_member_required  # only staff/admin can access
@never_cache
def admin_profile(request, pk=None):
    """
    If pk is provided: show that user's profile (admin viewing others).
    If pk is None: show the current admin's own profile.
    """
    admin_user = request.user if pk is None else get_object_or_404(CustomUser, pk=pk)
    return render(request, 'admin1180/admin_profile.html', {'admin_user': admin_user})


# =========================
# Admin Reports (Page Shell)
# =========================
@staff_member_required
@never_cache
def admin_reports(request):
    return render(request, 'admin1180/admin_reports.html')


# =========================
# Admin Reports Dashboard
# =========================
@staff_member_required
@never_cache
def reports_dashboard(request):
    """Dashboard view for reports with statistics"""
    User = get_user_model()

    # Calculate statistics
    total_debt = Debtor.objects.filter(is_delete=False).aggregate(
        total=Sum('total_debt'))['total'] or 0

    context = {
        'title': 'Reports Dashboard',
        'total_users': User.objects.count(),
        'total_debtors': Debtor.objects.filter(is_delete=False).count(),
        'total_transactions': Transaction.objects.count(),
        'total_debt': total_debt,
        'debtors': Debtor.objects.filter(is_delete=False).order_by('name'),
    }

    return render(request, 'admin/reports_dashboard.html', context)


# =========================
# Helpers for Admin Excel Exports
# =========================
def _create_excel_response(filename_prefix: str) -> HttpResponse:
    """Create HttpResponse for Excel file download"""
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename_prefix}_{timestamp}.xlsx"'
    return response


def _style_worksheet_header(ws, headers):
    """Apply styling to worksheet headers"""
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center')

    ws.append(headers)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold_font
        cell.alignment = center_alignment

        column_letter = get_column_letter(col_idx)
        header_text = headers[col_idx - 1]
        ws.column_dimensions[column_letter].width = max(15, len(str(header_text)) + 3)

    ws.freeze_panes = "A2"


def _convert_to_excel_format(value):
    """Convert Python values to Excel-compatible format"""
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (datetime, time)):
        if hasattr(value, 'tzinfo') and value.tzinfo is not None:
            value = timezone.make_naive(value, timezone.get_current_timezone())
        return value
    if isinstance(value, date):
        return value
    if isinstance(value, bool):
        return "Yes" if value else "No"
    return str(value)


# =========================
# Admin Exports: All Users
# =========================
@staff_member_required
@never_cache
def export_all_users_xlsx(request):
    """Export all users to Excel file"""
    User = get_user_model()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "All Users"

    headers = [
        "User ID", "Username", "First Name", "Last Name", "Email",
        "Mobile", "Address", "Date Joined", "Last Login", "Is Active",
        "Is Staff", "Is Superuser", "Total Debtors"
    ]

    _style_worksheet_header(worksheet, headers)

    users_queryset = (
        User.objects
        .annotate(debtor_count=Count('debtors'))
        .select_related()
        .order_by('id')
    )

    for user in users_queryset:
        row_data = [
            user.id, user.username,
            getattr(user, 'first_name', '') or '',
            getattr(user, 'last_name', '') or '',
            getattr(user, 'email', '') or '',
            getattr(user, 'mobile', '') or '',
            getattr(user, 'address', '') or '',
            user.date_joined, user.last_login,
            user.is_active, user.is_staff, user.is_superuser,
            user.debtor_count,
        ]

        excel_row = [_convert_to_excel_format(value) for value in row_data]
        worksheet.append(excel_row)

    response = _create_excel_response("all_users")
    workbook.save(response)
    return response


# =========================
# Admin Exports: All Debtors
# =========================
@staff_member_required
@never_cache
def export_all_debtors_xlsx(request):
    """Export all debtors to Excel file"""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "All Debtors"

    headers = [
        "Debtor ID", "Name", "Mobile", "Address", "Status",
        "Initial Debt", "Total Debt", "Debt Date", "Purpose",
        "Payment Method", "Voucher/Cheque No", "Created By",
        "Created At", "Updated At", "Is Active"
    ]

    _style_worksheet_header(worksheet, headers)

    debtors_queryset = (
        Debtor.objects
        .select_related('created_by')
        .filter(is_delete=False)
        .order_by('id')
    )

    for debtor in debtors_queryset:
        row_data = [
            debtor.debtor_id, debtor.name, debtor.mobile,
            debtor.address, debtor.debtor_status,
            debtor.initial_debt, debtor.total_debt,
            debtor.debt_date, debtor.debt_purpose,
            debtor.payment_method, debtor.voucher_cheque_no,
            debtor.created_by.username if debtor.created_by else '',
            debtor.created_at, debtor.updated_at,
            not debtor.is_delete,
        ]

        excel_row = [_convert_to_excel_format(value) for value in row_data]
        worksheet.append(excel_row)

    response = _create_excel_response("all_debtors")
    workbook.save(response)
    return response


# =========================
# Admin Exports: All Transactions
# =========================
@staff_member_required
@never_cache
def export_all_transactions_xlsx(request):
    """Export all transactions from all debtors"""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "All Transactions"

    headers = [
        "Transaction ID", "Debtor ID", "Debtor Name", "Transaction Type",
        "Debit Amount", "Credit Amount", "Transaction Amount", "Current Debt",
        "Description", "Payment Method", "Transaction Date", "Recorded By",
        "Updated At"
    ]

    _style_worksheet_header(worksheet, headers)

    transactions_queryset = (
        Transaction.objects
        .select_related('debtor', 'recorded_by')
        .order_by('-tran_date')
    )

    for transaction in transactions_queryset:
        row_data = [
            transaction.tran_id,
            transaction.debtor.debtor_id,
            transaction.debtor.name,
            transaction.tran_type,
            transaction.debit_amount,
            transaction.credit_amount,
            transaction.tran_amount,
            transaction.current_debt,
            transaction.tran_desc,
            transaction.tran_medium,
            transaction.tran_date,
            transaction.recorded_by.username if transaction.recorded_by else '',
            transaction.updated_at,
        ]

        excel_row = [_convert_to_excel_format(value) for value in row_data]
        worksheet.append(excel_row)

    # Add summary
    last_row = worksheet.max_row + 2
    worksheet[f'A{last_row}'] = f"Total Transactions: {transactions_queryset.count()}"
    worksheet[f'A{last_row}'].font = Font(bold=True)
    worksheet[f'A{last_row + 1}'] = f"Report Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"

    response = _create_excel_response("all_transactions")
    workbook.save(response)
    return response



    # Debtors Summary Sheet
    debtors_sheet = workbook.create_sheet("Debtors Summary")
    debtors_headers = [
        "Debtor Name", "Mobile", "Status", "Total Debt", "Created By", "Debt Date"
    ]

    _style_worksheet_header(debtors_sheet, debtors_headers)

    debtors_summary = (
        Debtor.objects
        .select_related('created_by')
        .filter(is_delete=False)
        .order_by('-total_debt', 'name')
    )

    for debtor in debtors_summary:
        row_data = [
            debtor.name, debtor.mobile, debtor.debtor_status,
            debtor.total_debt,
            debtor.created_by.username if debtor.created_by else '',
            debtor.debt_date,
        ]

        excel_row = [_convert_to_excel_format(value) for value in row_data]
        debtors_sheet.append(excel_row)

    response = _create_excel_response("summary_report")
    workbook.save(response)
    return response


# =========================
# Admin Export: Single Debtor Transactions
# =========================
@staff_member_required
@never_cache
def export_debtor_transactions_xlsx(request):
    """Export transactions for a specific debtor"""
    debtor_id = request.GET.get('debtor_id')

    if not debtor_id:
        # Return list of debtors for selection
        debtors = Debtor.objects.filter(is_delete=False).order_by('name')
        context = {
            'debtors': debtors,
            'title': 'Select Debtor for Transaction Export'
        }
        return render(request, 'admin/reports_dashboard.html', context)

    try:
        debtor = Debtor.objects.get(id=debtor_id, is_delete=False)
    except Debtor.DoesNotExist:
        return HttpResponse("Invalid debtor selected", status=400)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = f"Transactions - {debtor.name}"

    headers = [
        "Transaction ID", "Debtor Name", "Transaction Type", "Debit Amount",
        "Credit Amount", "Transaction Amount", "Current Debt After Transaction",
        "Description", "Payment Method", "Transaction Date", "Recorded By"
    ]

    _style_worksheet_header(worksheet, headers)

    # Add debtor info header
    worksheet.insert_rows(1)
    worksheet['A1'] = f"Transaction Report for: {debtor.name} (ID: {debtor.debtor_id})"
    worksheet['A1'].font = Font(bold=True, size=14)
    worksheet.merge_cells('A1:K1')

    transactions_queryset = (
        Transaction.objects
        .filter(debtor=debtor)
        .select_related('recorded_by', 'debtor')
        .order_by('-tran_date')
    )

    for transaction in transactions_queryset:
        row_data = [
            transaction.tran_id, debtor.name, transaction.tran_type,
            transaction.debit_amount, transaction.credit_amount,
            transaction.tran_amount, transaction.current_debt,
            transaction.tran_desc, transaction.tran_medium,
            transaction.tran_date,
            transaction.recorded_by.username if transaction.recorded_by else '',
        ]

        excel_row = [_convert_to_excel_format(value) for value in row_data]
        worksheet.append(excel_row)

    # Add summary
    last_row = worksheet.max_row + 2
    worksheet[f'A{last_row}'] = "Summary:"
    worksheet[f'A{last_row}'].font = Font(bold=True)
    worksheet[f'A{last_row + 1}'] = f"Total Transactions: {transactions_queryset.count()}"
    worksheet[f'A{last_row + 2}'] = f"Current Debt: {debtor.total_debt}"
    worksheet[f'A{last_row + 3}'] = f"Report Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"

    safe_name = "".join(c for c in debtor.name if c.isalnum() or c in (' ', '-', '_')).rstrip()
    response = _create_excel_response(f"transactions_{safe_name}")
    workbook.save(response)
    return response

##Terms & Conditon
def terms_condition(request):
    return render(request, 'terms_condition.html')